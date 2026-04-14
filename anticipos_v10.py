"""
Sistema de Registro y Legalización de Anticipos - Transporte de Carga
Colombia - Conectado a Supabase (PostgreSQL)
v10: + Vacaciones de conductores + Préstamos con trazabilidad
"""

import streamlit as st
import psycopg2
import pandas as pd
from datetime import datetime, timedelta, date
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==================== CONFIGURACIÓN ====================
SUPABASE_DB_URL = "postgresql://postgres.symbkeqadbevebtanodh:Conejito900@aws-1-us-east-1.pooler.supabase.com:6543/postgres"

# ==================== FORMATO COLOMBIANO ====================
def fmt(valor):
    if valor is None:
        return "0"
    try:
        return f"{int(float(valor)):,}".replace(',', '.')
    except:
        return str(valor)

def limpiar(texto):
    if not texto:
        return 0.0
    try:
        return float(str(texto).replace('.', '').replace(',', '.'))
    except:
        return 0.0

def hora_colombia():
    return datetime.utcnow() - timedelta(hours=5)

# ==================== ALERTAS ANTICIPOS ====================
def clasificar_alerta(fecha_viaje):
    hoy = hora_colombia().date()
    try:
        fv = fecha_viaje.date() if hasattr(fecha_viaje, 'date') else pd.to_datetime(fecha_viaje).date()
    except:
        return 0, "ok"
    dias = (hoy - fv).days
    if dias <= 3:
        return dias, "ok"
    elif dias <= 7:
        return dias, "warning"
    else:
        return dias, "critical"

def badge_alerta(dias, nivel):
    if nivel == "critical":
        return f"🔴 {dias}d"
    elif nivel == "warning":
        return f"🟡 {dias}d"
    else:
        return f"🟢 {dias}d"

# ==================== EXPORTAR EXCEL ANTICIPOS ====================
def generar_excel(df: pd.DataFrame, titulo: str = "Anticipos") -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Anticipos"
    color_header   = "1F4E79"
    color_critico  = "FCE4EC"
    color_warning  = "FFF9C4"
    color_ok       = "E8F5E9"
    color_leg      = "E3F2FD"
    color_subtotal = "BBDEFB"
    font_header = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    font_titulo = Font(name="Arial", bold=True, size=13, color="1F4E79")
    font_normal = Font(name="Arial", size=9)
    font_bold   = Font(name="Arial", bold=True, size=9)
    font_red    = Font(name="Arial", bold=True, size=9, color="C62828")
    font_subtot = Font(name="Arial", bold=True, size=10, color="1F4E79")
    thin   = Side(style="thin", color="BDBDBD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")
    ws.merge_cells("A1:M1")
    ws["A1"] = f"Reporte de Anticipos — {titulo}"
    ws["A1"].font = font_titulo
    ws["A1"].alignment = center
    ws.merge_cells("A2:M2")
    ws["A2"] = f"Generado: {hora_colombia().strftime('%d/%m/%Y %H:%M')} (hora Colombia)"
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="757575")
    ws["A2"].alignment = center
    columnas = ["ID","Manifiesto","Fecha viaje","Placa","Conductor","Cliente","Origen","Destino",
                "Anticipo (COP)","Estado","Días pend.","Legalizado por","Fecha legalización"]
    row_header = 4
    for col_idx, col_name in enumerate(columnas, start=1):
        cell = ws.cell(row=row_header, column=col_idx, value=col_name)
        cell.font = font_header
        cell.fill = PatternFill("solid", fgColor=color_header)
        cell.alignment = center
        cell.border = border
    for row_idx, (_, row) in enumerate(df.iterrows(), start=row_header + 1):
        legalizado = bool(row.get("legalizado", False))
        dias, nivel = clasificar_alerta(row.get("fecha_viaje"))
        if legalizado:
            row_color = color_leg
        elif nivel == "critical":
            row_color = color_critico
        elif nivel == "warning":
            row_color = color_warning
        else:
            row_color = color_ok
        fill = PatternFill("solid", fgColor=row_color)
        valores = [
            row.get("id",""), row.get("manifiesto",""),
            str(row.get("fecha_viaje",""))[:10], row.get("placa",""),
            row.get("conductor",""), row.get("cliente",""),
            row.get("origen",""), row.get("destino",""),
            int(row.get("valor_anticipo",0)),
            "Legalizado" if legalizado else "Pendiente",
            "" if legalizado else dias,
            row.get("legalizado_por","") or "",
            str(row.get("fecha_legalizacion","") or "")[:16],
        ]
        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.fill = fill; cell.border = border
            cell.alignment = center if col_idx in [1,10,11] else left
            if col_idx == 9 and not legalizado and nivel == "critical":
                cell.font = font_red
            elif col_idx == 9:
                cell.font = font_bold
            else:
                cell.font = font_normal
    total_row = row_header + len(df) + 2
    ws.cell(row=total_row, column=8, value="TOTAL ANTICIPOS:").font = font_subtot
    ws.cell(row=total_row, column=8).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=9, value=f'=SUM(I{row_header+1}:I{row_header+len(df)})').font = font_subtot
    ws.cell(row=total_row, column=9).fill = PatternFill("solid", fgColor=color_subtotal)
    ws.cell(row=total_row, column=9).border = border
    ws.cell(row=total_row, column=9).alignment = center
    anchos = [6,14,13,10,22,20,18,18,18,16,10,22,20]
    for col_idx, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho
    for row_idx in range(row_header+1, row_header+len(df)+1):
        ws.cell(row=row_idx, column=9).number_format = '#,##0'
    ws.freeze_panes = f"A{row_header + 1}"
    output = BytesIO(); wb.save(output); output.seek(0)
    return output

# ==================== EXPORTAR EXCEL PRÉSTAMOS ====================
def generar_excel_prestamos(df_prestamos: pd.DataFrame, df_pagos: pd.DataFrame) -> BytesIO:
    wb = Workbook()

    # Hoja 1: Resumen préstamos
    ws1 = wb.active
    ws1.title = "Préstamos"
    color_h    = "1F4E79"
    color_paz  = "E8F5E9"
    color_deu  = "FCE4EC"
    color_sub  = "BBDEFB"
    fh  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    fn  = Font(name="Arial", size=9)
    fb  = Font(name="Arial", bold=True, size=9)
    ft  = Font(name="Arial", bold=True, size=13, color="1F4E79")
    fs  = Font(name="Arial", bold=True, size=10, color="1F4E79")
    thin   = Side(style="thin", color="BDBDBD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left_a = Alignment(horizontal="left",   vertical="center")
    ws1.merge_cells("A1:I1")
    ws1["A1"] = f"Reporte de Préstamos — {hora_colombia().strftime('%d/%m/%Y %H:%M')}"
    ws1["A1"].font = ft; ws1["A1"].alignment = center
    cols_prest = ["ID","Conductor","Fecha préstamo","Monto total","Total pagado","Saldo pendiente","Estado","Motivo","Observaciones"]
    rh = 3
    for ci, cn in enumerate(cols_prest, 1):
        cell = ws1.cell(row=rh, column=ci, value=cn)
        cell.font = fh
        cell.fill = PatternFill("solid", fgColor=color_h)
        cell.alignment = center; cell.border = border
    for ri, (_, row) in enumerate(df_prestamos.iterrows(), start=rh+1):
        paz = row.get("estado","") == "saldado"
        fill = PatternFill("solid", fgColor=color_paz if paz else color_deu)
        pid = row.get("id",0)
        pagos_conductor = df_pagos[df_pagos["prestamo_id"]==pid]["monto_pago"].sum() if not df_pagos.empty else 0
        saldo = max(0, int(row.get("monto_total",0)) - int(pagos_conductor))
        valores = [
            pid, row.get("conductor",""),
            str(row.get("fecha_prestamo",""))[:10],
            int(row.get("monto_total",0)),
            int(pagos_conductor), saldo,
            "Paz y salvo" if paz else "Pendiente",
            row.get("motivo","") or "", row.get("observaciones","") or ""
        ]
        for ci, val in enumerate(valores, 1):
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.fill = fill; cell.border = border; cell.font = fn
            cell.alignment = center if ci in [1,4,5,6,7] else left_a
            if ci in [4,5,6]: cell.number_format = '#,##0'
    anchos1 = [6,22,14,16,16,16,14,22,28]
    for ci, aw in enumerate(anchos1, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = aw
    ws1.freeze_panes = f"A{rh+1}"

    # Hoja 2: Detalle pagos
    if not df_pagos.empty:
        ws2 = wb.create_sheet("Detalle pagos")
        ws2.merge_cells("A1:F1")
        ws2["A1"] = "Detalle de pagos / descuentos"
        ws2["A1"].font = ft; ws2["A1"].alignment = center
        cols_pago = ["ID pago","Préstamo ID","Conductor","Fecha pago","Monto descuento","Observaciones"]
        rh2 = 3
        for ci, cn in enumerate(cols_pago, 1):
            cell = ws2.cell(row=rh2, column=ci, value=cn)
            cell.font = fh
            cell.fill = PatternFill("solid", fgColor=color_h)
            cell.alignment = center; cell.border = border
        for ri, (_, row) in enumerate(df_pagos.iterrows(), start=rh2+1):
            fill2 = PatternFill("solid", fgColor="F3F3F3")
            pid2 = row.get("prestamo_id",0)
            cond2 = ""
            if not df_prestamos.empty:
                match = df_prestamos[df_prestamos["id"]==pid2]
                if not match.empty:
                    cond2 = match.iloc[0].get("conductor","")
            valores2 = [
                row.get("id",""), pid2, cond2,
                str(row.get("fecha_pago",""))[:10],
                int(row.get("monto_pago",0)),
                row.get("observaciones","") or ""
            ]
            for ci, val in enumerate(valores2, 1):
                cell = ws2.cell(row=ri, column=ci, value=val)
                cell.fill = fill2; cell.border = border; cell.font = fn
                cell.alignment = center if ci in [1,2,4,5] else left_a
                if ci == 5: cell.number_format = '#,##0'
        anchos2 = [8,12,22,14,18,30]
        for ci, aw in enumerate(anchos2, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = aw
        ws2.freeze_panes = f"A{rh2+1}"

    output = BytesIO(); wb.save(output); output.seek(0)
    return output

# ==================== EXPORTAR EXCEL VACACIONES ====================
def generar_excel_vacaciones(df_info: pd.DataFrame, df_vac: pd.DataFrame, conductores: list) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vacaciones"
    color_h   = "1F4E79"
    color_venc = "FCE4EC"
    color_prox = "FFF9C4"
    color_ok_  = "E8F5E9"
    fh = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    fn = Font(name="Arial", size=9)
    ft = Font(name="Arial", bold=True, size=13, color="1F4E79")
    thin   = Side(style="thin", color="BDBDBD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left_a = Alignment(horizontal="left",   vertical="center")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Reporte de Vacaciones — {hora_colombia().strftime('%d/%m/%Y %H:%M')}"
    ws["A1"].font = ft; ws["A1"].alignment = center
    cols = ["Conductor","Fecha ingreso","Días acumulados","Días tomados","Días pendientes","Prox. vacación","Estado"]
    rh = 3
    for ci, cn in enumerate(cols, 1):
        cell = ws.cell(row=rh, column=ci, value=cn)
        cell.font = fh
        cell.fill = PatternFill("solid", fgColor=color_h)
        cell.alignment = center; cell.border = border
    hoy = hora_colombia().date()
    for ri, cond in enumerate(sorted(conductores), start=rh+1):
        info_row = df_info[df_info["conductor"]==cond].iloc[0] if not df_info.empty and (df_info["conductor"]==cond).any() else None
        dias_tom = int(df_vac[df_vac["conductor"]==cond]["dias"].sum()) if not df_vac.empty and (df_vac["conductor"]==cond).any() else 0
        if info_row is not None and info_row.get("fecha_ingreso") is not None:
            fi = pd.to_datetime(info_row["fecha_ingreso"]).date()
            anios = (hoy - fi).days / 365.25
            dias_acum = int(anios * 15)
            dias_pend = max(0, dias_acum - dias_tom)
            prox_aniv = fi.replace(year=fi.year + int(anios) + 1)
            estado = "Vencidas" if dias_pend > 15 else ("Próximas" if dias_pend > 0 else "Al día")
            fill_color = color_venc if dias_pend > 15 else (color_prox if dias_pend > 0 else color_ok_)
        else:
            dias_acum = "—"; dias_pend = "—"; prox_aniv = "—"; estado = "Sin fecha ingreso"
            fill_color = "F3F3F3"
        fill = PatternFill("solid", fgColor=fill_color)
        fi_str = str(info_row["fecha_ingreso"])[:10] if info_row is not None and info_row.get("fecha_ingreso") is not None else "—"
        valores = [cond, fi_str, dias_acum, dias_tom, dias_pend, str(prox_aniv)[:10], estado]
        for ci, val in enumerate(valores, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill; cell.border = border; cell.font = fn
            cell.alignment = center if ci in [3,4,5,7] else left_a
    anchos = [28,14,16,14,16,16,18]
    for ci, aw in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(ci)].width = aw
    ws.freeze_panes = f"A{rh+1}"
    output = BytesIO(); wb.save(output); output.seek(0)
    return output

# ==================== PLACAS ====================
PLACAS = [
    "NOX459","NOX460","NOX461","SON047","SON048",
    "SOP148","SOP149","SOP150","SRO661","SRO672",
    "TMW882","TRL282","TRL298","UYQ308","UYV084",
    "UYY788","PSX350"
]

CONDUCTORES_DEFAULT = [
    "CARLOS TAFUR","CHRISTIAN MARTINEZ","EDGAR DE JESUS",
    "EDUARDO OLIVARES","FLAVIO MALTE","GONZALO","ISAIAS VESGA",
    "JOSE ORTEGA","JULIAN CALETH","PEDRO JR","RAMON TAFUR",
    "REIMUR MANUEL","SLITH ORTEGA","YEIMI DUQUE","SIN CONDUCTOR ASIGNADO",
]

CLIENTES_DEFAULT = [
    "GLOBO EXPRESS","MOTOTRANSPORTAMOS","CARGO ANDINA","TRANSOLICAR","SUCLOGISTIC",
]

# ==================== BASE DE DATOS ====================
class DB:
    def __init__(self):
        self.url = SUPABASE_DB_URL
        self._init_tablas()

    def conn(self):
        return psycopg2.connect(self.url)

    def _init_tablas(self):
        try:
            c = self.conn(); cur = c.cursor()
            # Tabla principal anticipos
            cur.execute("""
                CREATE TABLE IF NOT EXISTS anticipos_v1 (
                    id SERIAL PRIMARY KEY,
                    fecha_viaje DATE NOT NULL,
                    fecha_registro TIMESTAMP NOT NULL,
                    placa TEXT NOT NULL,
                    conductor TEXT NOT NULL,
                    cliente TEXT NOT NULL,
                    origen TEXT NOT NULL,
                    destino TEXT NOT NULL,
                    valor_anticipo BIGINT NOT NULL,
                    observaciones TEXT,
                    legalizado BOOLEAN DEFAULT FALSE,
                    fecha_legalizacion TIMESTAMP,
                    legalizado_por TEXT,
                    obs_legalizacion TEXT
                )
            """)
            cur.execute("ALTER TABLE anticipos_v1 ADD COLUMN IF NOT EXISTS manifiesto TEXT DEFAULT ''")
            # Catálogos
            cur.execute("""
                CREATE TABLE IF NOT EXISTS clientes_extra (
                    id SERIAL PRIMARY KEY,
                    nombre TEXT UNIQUE NOT NULL,
                    fecha_registro TIMESTAMP NOT NULL
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS conductores_extra (
                    id SERIAL PRIMARY KEY,
                    nombre TEXT UNIQUE NOT NULL,
                    fecha_registro TIMESTAMP NOT NULL
                )
            """)
            # Información de conductores (fecha ingreso, etc.)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS conductores_info (
                    id SERIAL PRIMARY KEY,
                    conductor TEXT UNIQUE NOT NULL,
                    fecha_ingreso DATE,
                    observaciones TEXT,
                    fecha_registro TIMESTAMP NOT NULL
                )
            """)
            # Vacaciones
            cur.execute("""
                CREATE TABLE IF NOT EXISTS vacaciones (
                    id SERIAL PRIMARY KEY,
                    conductor TEXT NOT NULL,
                    fecha_inicio DATE NOT NULL,
                    fecha_fin DATE NOT NULL,
                    dias INTEGER NOT NULL,
                    observaciones TEXT,
                    registrado_por TEXT,
                    fecha_registro TIMESTAMP NOT NULL
                )
            """)
            # Préstamos
            cur.execute("""
                CREATE TABLE IF NOT EXISTS prestamos (
                    id SERIAL PRIMARY KEY,
                    conductor TEXT NOT NULL,
                    monto_total BIGINT NOT NULL,
                    fecha_prestamo DATE NOT NULL,
                    motivo TEXT,
                    observaciones TEXT,
                    estado TEXT DEFAULT 'activo',
                    fecha_registro TIMESTAMP NOT NULL
                )
            """)
            # Pagos / descuentos de préstamos
            cur.execute("""
                CREATE TABLE IF NOT EXISTS pagos_prestamos (
                    id SERIAL PRIMARY KEY,
                    prestamo_id INTEGER NOT NULL REFERENCES prestamos(id) ON DELETE CASCADE,
                    monto_pago BIGINT NOT NULL,
                    fecha_pago DATE NOT NULL,
                    observaciones TEXT,
                    registrado_por TEXT,
                    fecha_registro TIMESTAMP NOT NULL
                )
            """)
            c.commit(); c.close()
        except Exception as e:
            st.error(f"Error inicializando tablas: {e}")

    # ---- Clientes ----
    def obtener_clientes_extra(self):
        try:
            c = self.conn()
            df = pd.read_sql_query("SELECT * FROM clientes_extra ORDER BY nombre", c)
            c.close(); return df
        except:
            return pd.DataFrame(columns=['id','nombre','fecha_registro'])

    def agregar_cliente(self, nombre):
        try:
            c = self.conn(); cur = c.cursor()
            cur.execute("INSERT INTO clientes_extra (nombre, fecha_registro) VALUES (%s, %s)",
                        (nombre.strip().upper(), hora_colombia()))
            c.commit(); c.close(); return True
        except:
            return False

    def eliminar_cliente(self, cliente_id):
        try:
            c = self.conn(); cur = c.cursor()
            cur.execute("DELETE FROM clientes_extra WHERE id = %s", (cliente_id,))
            c.commit(); c.close()
        except Exception as e:
            st.error(f"Error eliminando cliente: {e}")

    # ---- Conductores extra ----
    def obtener_conductores_extra(self):
        try:
            c = self.conn()
            df = pd.read_sql_query("SELECT * FROM conductores_extra ORDER BY nombre", c)
            c.close(); return df
        except:
            return pd.DataFrame(columns=['id','nombre','fecha_registro'])

    def agregar_conductor(self, nombre):
        try:
            c = self.conn(); cur = c.cursor()
            cur.execute("INSERT INTO conductores_extra (nombre, fecha_registro) VALUES (%s, %s)",
                        (nombre.strip().upper(), hora_colombia()))
            c.commit(); c.close(); return True
        except:
            return False

    def editar_conductor(self, conductor_id, nombre_nuevo):
        try:
            c = self.conn(); cur = c.cursor()
            cur.execute("UPDATE conductores_extra SET nombre = %s WHERE id = %s",
                        (nombre_nuevo.strip().upper(), conductor_id))
            c.commit(); c.close(); return True
        except Exception as e:
            st.error(f"Error editando conductor: {e}"); return False

    def eliminar_conductor(self, conductor_id):
        try:
            c = self.conn(); cur = c.cursor()
            cur.execute("DELETE FROM conductores_extra WHERE id = %s", (conductor_id,))
            c.commit(); c.close()
        except Exception as e:
            st.error(f"Error eliminando conductor: {e}")

    # ---- Conductores info ----
    def obtener_info_conductor(self, conductor):
        try:
            c = self.conn()
            df = pd.read_sql_query("SELECT * FROM conductores_info WHERE conductor = %s", c, params=(conductor,))
            c.close()
            return df.iloc[0] if not df.empty else None
        except:
            return None

    def obtener_todos_info_conductores(self):
        try:
            c = self.conn()
            df = pd.read_sql_query("SELECT * FROM conductores_info ORDER BY conductor", c)
            c.close(); return df
        except:
            return pd.DataFrame(columns=['id','conductor','fecha_ingreso','observaciones','fecha_registro'])

    def guardar_info_conductor(self, conductor, fecha_ingreso, observaciones=""):
        try:
            c = self.conn(); cur = c.cursor()
            cur.execute("""
                INSERT INTO conductores_info (conductor, fecha_ingreso, observaciones, fecha_registro)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (conductor) DO UPDATE
                SET fecha_ingreso = EXCLUDED.fecha_ingreso,
                    observaciones = EXCLUDED.observaciones
            """, (conductor.strip().upper(), fecha_ingreso, observaciones.strip(), hora_colombia()))
            c.commit(); c.close(); return True
        except Exception as e:
            st.error(f"Error guardando info conductor: {e}"); return False

    # ---- Vacaciones ----
    def obtener_vacaciones(self, conductor=None):
        try:
            c = self.conn()
            if conductor:
                df = pd.read_sql_query(
                    "SELECT * FROM vacaciones WHERE conductor = %s ORDER BY fecha_inicio DESC",
                    c, params=(conductor,))
            else:
                df = pd.read_sql_query("SELECT * FROM vacaciones ORDER BY fecha_inicio DESC", c)
            c.close(); return df
        except:
            return pd.DataFrame()

    def registrar_vacacion(self, data):
        try:
            c = self.conn(); cur = c.cursor()
            cur.execute("""
                INSERT INTO vacaciones (conductor, fecha_inicio, fecha_fin, dias, observaciones, registrado_por, fecha_registro)
                VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id
            """, (
                data['conductor'], data['fecha_inicio'], data['fecha_fin'],
                int(data['dias']), data.get('observaciones',''),
                data.get('registrado_por','').strip().upper(), hora_colombia()
            ))
            nuevo_id = cur.fetchone()[0]
            c.commit(); c.close(); return nuevo_id
        except Exception as e:
            st.error(f"Error registrando vacación: {e}"); return None

    def eliminar_vacacion(self, vac_id):
        try:
            c = self.conn(); cur = c.cursor()
            cur.execute("DELETE FROM vacaciones WHERE id = %s", (vac_id,))
            c.commit(); c.close()
        except Exception as e:
            st.error(f"Error eliminando vacación: {e}")

    # ---- Préstamos ----
    def obtener_prestamos(self, conductor=None, estado=None):
        try:
            c = self.conn()
            q = "SELECT * FROM prestamos WHERE 1=1"
            params = []
            if conductor:
                q += " AND conductor = %s"; params.append(conductor)
            if estado and estado != "Todos":
                q += " AND estado = %s"; params.append(estado)
            q += " ORDER BY fecha_prestamo DESC, fecha_registro DESC"
            df = pd.read_sql_query(q, c, params=params or None)
            c.close(); return df
        except Exception as e:
            st.error(f"Error obteniendo préstamos: {e}"); return pd.DataFrame()

    def registrar_prestamo(self, data):
        try:
            c = self.conn(); cur = c.cursor()
            cur.execute("""
                INSERT INTO prestamos (conductor, monto_total, fecha_prestamo, motivo, observaciones, estado, fecha_registro)
                VALUES (%s, %s, %s, %s, %s, 'activo', %s) RETURNING id
            """, (
                data['conductor'], int(data['monto_total']),
                data['fecha_prestamo'], data.get('motivo','').strip(),
                data.get('observaciones','').strip(), hora_colombia()
            ))
            nuevo_id = cur.fetchone()[0]
            c.commit(); c.close(); return nuevo_id
        except Exception as e:
            st.error(f"Error registrando préstamo: {e}"); return None

    def eliminar_prestamo(self, prestamo_id):
        try:
            c = self.conn(); cur = c.cursor()
            cur.execute("DELETE FROM prestamos WHERE id = %s", (prestamo_id,))
            c.commit(); c.close()
        except Exception as e:
            st.error(f"Error eliminando préstamo: {e}")

    def actualizar_estado_prestamo(self, prestamo_id, estado):
        try:
            c = self.conn(); cur = c.cursor()
            cur.execute("UPDATE prestamos SET estado = %s WHERE id = %s", (estado, prestamo_id))
            c.commit(); c.close()
        except Exception as e:
            st.error(f"Error actualizando estado: {e}")

    # ---- Pagos ----
    def obtener_pagos(self, prestamo_id=None):
        try:
            c = self.conn()
            if prestamo_id:
                df = pd.read_sql_query(
                    "SELECT * FROM pagos_prestamos WHERE prestamo_id = %s ORDER BY fecha_pago DESC",
                    c, params=(prestamo_id,))
            else:
                df = pd.read_sql_query("SELECT * FROM pagos_prestamos ORDER BY fecha_pago DESC", c)
            c.close(); return df
        except:
            return pd.DataFrame()

    def registrar_pago(self, data):
        try:
            c = self.conn(); cur = c.cursor()
            cur.execute("""
                INSERT INTO pagos_prestamos (prestamo_id, monto_pago, fecha_pago, observaciones, registrado_por, fecha_registro)
                VALUES (%s, %s, %s, %s, %s, %s) RETURNING id
            """, (
                data['prestamo_id'], int(data['monto_pago']),
                data['fecha_pago'], data.get('observaciones','').strip(),
                data.get('registrado_por','').strip().upper(), hora_colombia()
            ))
            nuevo_id = cur.fetchone()[0]
            c.commit(); c.close(); return nuevo_id
        except Exception as e:
            st.error(f"Error registrando pago: {e}"); return None

    def eliminar_pago(self, pago_id):
        try:
            c = self.conn(); cur = c.cursor()
            cur.execute("DELETE FROM pagos_prestamos WHERE id = %s", (pago_id,))
            c.commit(); c.close()
        except Exception as e:
            st.error(f"Error eliminando pago: {e}")

    # ---- Anticipos ----
    def registrar_viaje(self, data):
        try:
            c = self.conn(); cur = c.cursor()
            cur.execute("""
                INSERT INTO anticipos_v1
                (fecha_viaje, fecha_registro, placa, conductor, cliente,
                 origen, destino, valor_anticipo, observaciones, manifiesto, legalizado)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, FALSE) RETURNING id
            """, (
                data['fecha_viaje'], hora_colombia(), data['placa'], data['conductor'],
                data['cliente'], data['origen'], data['destino'],
                int(data['valor_anticipo']), data.get('observaciones',''),
                data.get('manifiesto','').strip().upper()
            ))
            nuevo_id = cur.fetchone()[0]
            c.commit(); c.close(); return nuevo_id
        except Exception as e:
            st.error(f"Error guardando viaje: {e}"); return None

    def editar_viaje(self, viaje_id, data):
        try:
            c = self.conn(); cur = c.cursor()
            cur.execute("""
                UPDATE anticipos_v1 SET
                    fecha_viaje = %s, placa = %s, conductor = %s, cliente = %s,
                    origen = %s, destino = %s, valor_anticipo = %s,
                    observaciones = %s, manifiesto = %s
                WHERE id = %s
            """, (
                data['fecha_viaje'], data['placa'], data['conductor'], data['cliente'],
                data['origen'], data['destino'], int(data['valor_anticipo']),
                data.get('observaciones',''), data.get('manifiesto','').strip().upper(), viaje_id
            ))
            c.commit(); c.close(); return True
        except Exception as e:
            st.error(f"Error editando viaje: {e}"); return False

    def legalizar(self, viaje_id, nombre_quien_legaliza, obs_legalizacion=""):
        try:
            c = self.conn(); cur = c.cursor()
            cur.execute("""
                UPDATE anticipos_v1
                SET legalizado = TRUE, fecha_legalizacion = %s,
                    legalizado_por = %s, obs_legalizacion = %s
                WHERE id = %s
            """, (hora_colombia(), nombre_quien_legaliza, obs_legalizacion, viaje_id))
            c.commit(); c.close(); return True
        except Exception as e:
            st.error(f"Error legalizando: {e}"); return False

    def buscar(self, estado=None, fecha_ini=None, fecha_fin=None, placa=None, conductor=None, manifiesto=None):
        try:
            c = self.conn()
            q = "SELECT * FROM anticipos_v1 WHERE 1=1"; params = []
            if estado == "legalizado":   q += " AND legalizado = TRUE"
            elif estado == "pendiente":  q += " AND legalizado = FALSE"
            if fecha_ini: q += " AND fecha_viaje >= %s"; params.append(fecha_ini)
            if fecha_fin: q += " AND fecha_viaje <= %s"; params.append(fecha_fin)
            if placa:     q += " AND placa = %s";        params.append(placa)
            if conductor: q += " AND conductor ILIKE %s"; params.append(f"%{conductor}%")
            if manifiesto:q += " AND manifiesto ILIKE %s"; params.append(f"%{manifiesto}%")
            q += " ORDER BY fecha_viaje DESC, fecha_registro DESC"
            df = pd.read_sql_query(q, c, params=params); c.close(); return df
        except Exception as e:
            st.error(f"Error buscando: {e}"); return pd.DataFrame()

    def eliminar(self, viaje_id):
        try:
            c = self.conn(); cur = c.cursor()
            cur.execute("DELETE FROM anticipos_v1 WHERE id = %s", (viaje_id,))
            c.commit(); c.close()
        except Exception as e:
            st.error(f"Error eliminando: {e}")

    def obtener_por_id(self, viaje_id):
        try:
            c = self.conn()
            df = pd.read_sql_query("SELECT * FROM anticipos_v1 WHERE id = %s", c, params=(viaje_id,))
            c.close()
            return df.iloc[0] if not df.empty else None
        except:
            return None


# ==================== HELPERS ====================
def get_lista_clientes(db):
    extras_df = db.obtener_clientes_extra()
    extras = extras_df['nombre'].tolist() if not extras_df.empty else []
    return sorted(set(CLIENTES_DEFAULT + extras))

def get_lista_conductores(db):
    extras_df = db.obtener_conductores_extra()
    extras = extras_df['nombre'].tolist() if not extras_df.empty else []
    return sorted(set(CONDUCTORES_DEFAULT + extras))

def calcular_vacaciones(conductor, df_info, df_vac):
    """Retorna dict con días acumulados, tomados, pendientes, próxima vacación"""
    hoy = hora_colombia().date()
    info = df_info[df_info["conductor"] == conductor].iloc[0] if not df_info.empty and (df_info["conductor"] == conductor).any() else None
    dias_tomados = int(df_vac[df_vac["conductor"] == conductor]["dias"].sum()) if not df_vac.empty and (df_vac["conductor"] == conductor).any() else 0

    if info is not None and info.get("fecha_ingreso") is not None:
        fi = pd.to_datetime(info["fecha_ingreso"]).date()
        anios_trabajados = (hoy - fi).days / 365.25
        dias_acum = int(anios_trabajados * 15)
        dias_pend = max(0, dias_acum - dias_tomados)
        # Próximo aniversario
        anio_actual_aniv = fi.year + int(anios_trabajados)
        try:
            prox_aniv = fi.replace(year=anio_actual_aniv + 1)
        except ValueError:
            prox_aniv = date(anio_actual_aniv + 1, fi.month, 28)
        dias_para_prox = (prox_aniv - hoy).days
        return {
            "fecha_ingreso": fi,
            "anios": round(anios_trabajados, 1),
            "dias_acum": dias_acum,
            "dias_tomados": dias_tomados,
            "dias_pend": dias_pend,
            "prox_aniv": prox_aniv,
            "dias_para_prox": dias_para_prox,
        }
    else:
        return {
            "fecha_ingreso": None,
            "anios": None,
            "dias_acum": None,
            "dias_tomados": dias_tomados,
            "dias_pend": None,
            "prox_aniv": None,
            "dias_para_prox": None,
        }

def calcular_saldo_prestamo(prestamo_id, monto_total, df_pagos):
    """Retorna (pagado, saldo)"""
    if df_pagos.empty:
        return 0, int(monto_total)
    pagos_p = df_pagos[df_pagos["prestamo_id"] == prestamo_id]
    pagado = int(pagos_p["monto_pago"].sum())
    saldo  = max(0, int(monto_total) - pagado)
    return pagado, saldo


# ==================== APP PRINCIPAL ====================
def main():
    st.set_page_config(page_title="Anticipos - Transporte de Carga", layout="wide", page_icon="🚛")
    st.title("🚛 Gestión de Anticipos - Transporte de Carga")

    # Session state
    for key, val in [
        ('db', None), ('confirmar_eliminar', None), ('editando_id', None),
        ('confirmar_eliminar_cliente', None), ('confirmar_eliminar_conductor', None),
        ('editando_conductor_id', None), ('confirmar_eliminar_vac', None),
        ('confirmar_eliminar_prestamo', None), ('confirmar_eliminar_pago', None),
    ]:
        if key not in st.session_state:
            st.session_state[key] = val

    if st.session_state.db is None:
        st.session_state.db = DB()
    db = st.session_state.db

    (tab_reg, tab_leg, tab_hist,
     tab_vac, tab_prest,
     tab_clientes, tab_conductores) = st.tabs([
        "📝 Registrar Viaje",
        "✅ Legalizar Anticipos",
        "📋 Historial",
        "🏖️ Vacaciones",
        "💰 Préstamos",
        "🏢 Clientes",
        "👤 Conductores",
    ])

    # ==================== TAB 1: REGISTRAR ====================
    with tab_reg:
        st.header("Registrar nuevo viaje con anticipo")
        lista_clientes    = get_lista_clientes(db)
        lista_conductores = get_lista_conductores(db)

        with st.form("form_registro", clear_on_submit=True):
            col1, col2 = st.columns(2)
            with col1:
                fecha_viaje = st.date_input("Fecha del viaje", value=datetime.today())
                placa       = st.selectbox("Placa de la tractomula", PLACAS)
                conductor   = st.selectbox("Conductor", lista_conductores)
                cliente     = st.selectbox("Cliente", lista_clientes,
                                           help="Si no aparece, agrégalo en 🏢 Clientes")
            with col2:
                manifiesto  = st.text_input("Número de manifiesto ✱", placeholder="Ej: 1234567")
                origen      = st.text_input("Origen",  placeholder="Ciudad de origen")
                destino     = st.text_input("Destino", placeholder="Ciudad de destino")
                anticipo_txt = st.text_input("Valor del anticipo (COP)", placeholder="Ejemplo: 1.500.000")
                anticipo = limpiar(anticipo_txt)
                if anticipo > 0:
                    st.caption(f"💵 {fmt(anticipo)} COP")
                observaciones = st.text_area("Observaciones", height=80)

            submitted = st.form_submit_button("💾 Registrar Viaje", type="primary")
            if submitted:
                errores = []
                if not manifiesto.strip(): errores.append("El número de manifiesto es obligatorio")
                if not origen.strip():     errores.append("Origen es obligatorio")
                if not destino.strip():    errores.append("Destino es obligatorio")
                if anticipo <= 0:          errores.append("El valor del anticipo debe ser mayor a 0")
                if errores:
                    for e in errores: st.error(f"⚠️ {e}")
                else:
                    nuevo_id = db.registrar_viaje({
                        'fecha_viaje': fecha_viaje, 'placa': placa,
                        'conductor': conductor.strip().upper(),
                        'cliente': cliente.strip().upper(),
                        'origen': origen.strip().upper(),
                        'destino': destino.strip().upper(),
                        'valor_anticipo': anticipo,
                        'observaciones': observaciones.strip(),
                        'manifiesto': manifiesto.strip()
                    })
                    if nuevo_id:
                        st.success(f"""
✅ **Viaje registrado exitosamente (ID: {nuevo_id})**
- Manifiesto: **{manifiesto.strip().upper()}**
- Placa: {placa} | Conductor: {conductor.upper()}
- Ruta: {origen.upper()} → {destino.upper()}
- Cliente: {cliente.upper()}
- Anticipo: **${fmt(anticipo)} COP**
- Estado: 🔴 Pendiente de legalización
                        """)

    # ==================== TAB 2: LEGALIZAR ====================
    with tab_leg:
        st.header("Legalizar anticipos pendientes")
        col_f1, col_f2, col_f3, col_f4 = st.columns(4)
        with col_f1: fecha_ini_leg = st.date_input("Desde", value=None, key="leg_fi")
        with col_f2: fecha_fin_leg = st.date_input("Hasta", value=None, key="leg_ff")
        with col_f3: placa_leg = st.selectbox("Placa", ["Todas"] + PLACAS, key="leg_placa")
        with col_f4: manifiesto_leg = st.text_input("Buscar por manifiesto", placeholder="Nº manifiesto...", key="leg_manif")

        fi = fecha_ini_leg.strftime('%Y-%m-%d') if fecha_ini_leg else None
        ff = fecha_fin_leg.strftime('%Y-%m-%d') if fecha_fin_leg else None
        pl = None if placa_leg == "Todas" else placa_leg
        mf = manifiesto_leg.strip() if manifiesto_leg else None
        df_pendientes = db.buscar(estado="pendiente", fecha_ini=fi, fecha_fin=ff, placa=pl, manifiesto=mf)

        if df_pendientes.empty:
            st.success("✅ No hay anticipos pendientes de legalización.")
        else:
            criticos, atencion, al_dia = [], [], []
            for _, row in df_pendientes.iterrows():
                dias, nivel = clasificar_alerta(row['fecha_viaje'])
                entry = (row['id'], dias)
                if nivel == "critical": criticos.append(entry)
                elif nivel == "warning": atencion.append(entry)
                else: al_dia.append(entry)
            total_pendiente = df_pendientes['valor_anticipo'].sum()
            if criticos:
                st.error(f"🚨 **{len(criticos)} anticipo(s) CRÍTICO(S)** con más de 7 días sin legalizar  |  🟡 {len(atencion)} en atención  |  🟢 {len(al_dia)} al día  |  💰 Total: **${fmt(total_pendiente)} COP**")
            elif atencion:
                st.warning(f"🟡 **{len(atencion)} anticipo(s)** requieren atención (4-7 días)  |  🟢 {len(al_dia)} al día  |  💰 Total: **${fmt(total_pendiente)} COP**")
            else:
                st.info(f"🟢 {len(al_dia)} viaje(s) pendiente(s), todos al día  |  💰 Total: **${fmt(total_pendiente)} COP**")

            df_ordenado = df_pendientes.sort_values("fecha_viaje", ascending=False)
            for _, row in df_ordenado.iterrows():
                dias, nivel = clasificar_alerta(row['fecha_viaje'])
                badge = badge_alerta(dias, nivel)
                manif_label = f"Manif: {row.get('manifiesto','—')} | " if row.get('manifiesto') else ""
                label_expander = (f"{badge} | ID {row['id']} | {manif_label}"
                    f"{row['fecha_viaje']} | {row['placa']} | {row['conductor']} | "
                    f"{row['origen']} → {row['destino']} | ${fmt(row['valor_anticipo'])} COP")
                with st.expander(label_expander):
                    col_info, col_form = st.columns([2, 2])
                    with col_info:
                        st.markdown("**Datos del viaje:**")
                        if nivel == "critical": st.error(f"⏰ Este anticipo lleva **{dias} días** sin legalizar — acción urgente")
                        elif nivel == "warning": st.warning(f"⚠️ Este anticipo lleva **{dias} días** sin legalizar")
                        else: st.success(f"✅ {dias} días desde el viaje — al día")
                        st.write(f"📄 Manifiesto: **{row.get('manifiesto', '—')}**")
                        st.write(f"📅 Fecha: {row['fecha_viaje']}")
                        st.write(f"🚛 Placa: {row['placa']}")
                        st.write(f"👤 Conductor: {row['conductor']}")
                        st.write(f"🏢 Cliente: {row['cliente']}")
                        st.write(f"📍 Ruta: {row['origen']} → {row['destino']}")
                        st.write(f"💰 Anticipo: **${fmt(row['valor_anticipo'])} COP**")
                        if row.get('observaciones'): st.write(f"📝 Obs: {row['observaciones']}")
                    with col_form:
                        st.markdown("**Legalizar este viaje:**")
                        nombre_leg = st.text_input("Tu nombre completo (obligatorio)", placeholder="Escribe tu nombre...", key=f"nombre_leg_{row['id']}")
                        obs_leg = st.text_area("Observaciones (opcional)", height=80, key=f"obs_leg_{row['id']}")
                        if st.button("✅ Marcar como LEGALIZADO", key=f"btn_leg_{row['id']}", type="primary"):
                            if not nombre_leg.strip():
                                st.error("⚠️ Debes escribir tu nombre para poder legalizar.")
                            else:
                                ok = db.legalizar(row['id'], nombre_leg.strip().upper(), obs_leg.strip())
                                if ok:
                                    st.success(f"✅ Viaje ID {row['id']} legalizado por **{nombre_leg.upper()}** a las {hora_colombia().strftime('%H:%M')}")
                                    st.rerun()

    # ==================== TAB 3: HISTORIAL ====================
    with tab_hist:
        st.header("Historial de viajes")
        col1, col2, col3 = st.columns(3)
        with col1: estado_filtro = st.selectbox("Estado", ["Todos","Pendientes","Legalizados"], key="hist_estado")
        with col2: fecha_ini_h  = st.date_input("Desde", value=None, key="hist_fi")
        with col3: fecha_fin_h  = st.date_input("Hasta", value=None, key="hist_ff")
        col4, col5, col6 = st.columns(3)
        with col4: placa_h      = st.selectbox("Placa", ["Todas"] + PLACAS, key="hist_placa")
        with col5: conductor_h  = st.text_input("Buscar conductor", placeholder="Nombre parcial...", key="hist_cond")
        with col6: manifiesto_h = st.text_input("Buscar por manifiesto", placeholder="Nº manifiesto...", key="hist_manif")

        estado_map = {"Todos": None, "Pendientes": "pendiente", "Legalizados": "legalizado"}
        fi_h  = fecha_ini_h.strftime('%Y-%m-%d') if fecha_ini_h else None
        ff_h  = fecha_fin_h.strftime('%Y-%m-%d') if fecha_fin_h else None
        pl_h  = None if placa_h == "Todas" else placa_h
        df_hist = db.buscar(estado=estado_map[estado_filtro], fecha_ini=fi_h, fecha_fin=ff_h,
                            placa=pl_h, conductor=conductor_h or None, manifiesto=manifiesto_h.strip() or None)

        if df_hist.empty:
            st.info("No se encontraron viajes con los filtros aplicados.")
        else:
            total_anticipo = df_hist['valor_anticipo'].sum()
            legalizados = int(df_hist['legalizado'].sum())
            pendientes  = len(df_hist) - legalizados
            col_m1, col_m2, col_m3, col_m4 = st.columns(4)
            col_m1.metric("Total viajes",    len(df_hist))
            col_m2.metric("Legalizados",     legalizados)
            col_m3.metric("Pendientes",      pendientes)
            col_m4.metric("Total anticipos", f"${fmt(total_anticipo)}")

            if estado_filtro in ["Todos","Pendientes"]:
                criticos_hist = sum(1 for _, r in df_hist[df_hist['legalizado']==False].iterrows()
                                    if clasificar_alerta(r['fecha_viaje'])[1] == "critical")
                if criticos_hist > 0:
                    st.error(f"🚨 Hay **{criticos_hist} anticipo(s) crítico(s)** con más de 7 días sin legalizar.")

            cols_tabla = ['id','manifiesto','fecha_viaje','placa','conductor','cliente','origen',
                          'destino','valor_anticipo','legalizado','legalizado_por','fecha_legalizacion']
            df_show = df_hist[[c for c in cols_tabla if c in df_hist.columns]].copy()
            df_show['dias_alerta'] = df_hist.apply(
                lambda r: "—" if r.get('legalizado') else badge_alerta(*clasificar_alerta(r['fecha_viaje'])), axis=1)
            df_show['valor_anticipo'] = df_show['valor_anticipo'].apply(lambda x: f"${fmt(x)}")
            df_show['legalizado'] = df_show['legalizado'].apply(lambda x: "✅ Legalizado" if x else "🔴 Pendiente")
            df_show.rename(columns={
                'id':'ID','manifiesto':'Manifiesto','fecha_viaje':'Fecha viaje','placa':'Placa',
                'conductor':'Conductor','cliente':'Cliente','origen':'Origen','destino':'Destino',
                'valor_anticipo':'Anticipo','legalizado':'Estado','legalizado_por':'Legalizado por',
                'fecha_legalizacion':'Fecha legalización','dias_alerta':'Alerta'
            }, inplace=True)
            st.dataframe(df_show, use_container_width=True, hide_index=True, height=350)

            st.divider()
            col_exp1, col_exp2 = st.columns([3, 1])
            with col_exp1:
                titulo_excel = st.text_input("Título del reporte Excel",
                    value=f"Anticipos {estado_filtro} — {hora_colombia().strftime('%d/%m/%Y')}", key="titulo_excel")
            with col_exp2:
                st.markdown("&nbsp;")
                excel_bytes = generar_excel(df_hist, titulo_excel)
                st.download_button(label="📥 Exportar a Excel", data=excel_bytes,
                    file_name=f"anticipos_{hora_colombia().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")

            st.divider()
            st.subheader("Acciones sobre un viaje")
            viaje_sel = st.selectbox("Selecciona un viaje por ID", df_hist['id'].tolist(),
                format_func=lambda x: (
                    f"ID {x} | Manif: {df_hist[df_hist['id']==x]['manifiesto'].values[0] or '—'} | "
                    f"{df_hist[df_hist['id']==x]['placa'].values[0]} | {df_hist[df_hist['id']==x]['conductor'].values[0]}"
                ), key="hist_sel")
            row_sel = df_hist[df_hist['id'] == viaje_sel].iloc[0]
            col_det, col_acc = st.columns([3, 1])
            with col_det:
                estado_tag = "✅ **LEGALIZADO**" if row_sel['legalizado'] else "🔴 **PENDIENTE**"
                st.markdown(f"**Estado:** {estado_tag}")
                if not row_sel['legalizado']:
                    dias_sel, nivel_sel = clasificar_alerta(row_sel['fecha_viaje'])
                    if nivel_sel == "critical": st.error(f"⏰ Este anticipo lleva **{dias_sel} días** sin legalizar")
                    elif nivel_sel == "warning": st.warning(f"⚠️ {dias_sel} días sin legalizar")
                st.write(f"📄 Manifiesto: **{row_sel.get('manifiesto', '—')}**")
                st.write(f"Fecha: {row_sel['fecha_viaje']} | Placa: {row_sel['placa']} | Conductor: {row_sel['conductor']}")
                st.write(f"Cliente: {row_sel['cliente']}")
                st.write(f"Ruta: {row_sel['origen']} → {row_sel['destino']}")
                st.write(f"Anticipo: **${fmt(row_sel['valor_anticipo'])} COP**")
                if row_sel['legalizado']:
                    st.success(f"Legalizado por: **{row_sel['legalizado_por']}** | Fecha: {row_sel['fecha_legalizacion']}")
            with col_acc:
                st.markdown("&nbsp;")
                if st.button("✏️ Editar viaje", key="btn_editar"):
                    st.session_state.editando_id = viaje_sel; st.rerun()
                st.markdown("&nbsp;")
                if st.session_state.confirmar_eliminar == viaje_sel:
                    st.warning(f"¿Seguro que quieres eliminar ID **{viaje_sel}**?")
                    c_si, c_no = st.columns(2)
                    with c_si:
                        if st.button("Sí, eliminar", key="btn_si_eliminar", type="primary"):
                            db.eliminar(viaje_sel); st.session_state.confirmar_eliminar = None
                            st.success(f"Viaje ID {viaje_sel} eliminado."); st.rerun()
                    with c_no:
                        if st.button("Cancelar", key="btn_no_eliminar"):
                            st.session_state.confirmar_eliminar = None; st.rerun()
                else:
                    if st.button("🗑️ Eliminar viaje", key="btn_eliminar", type="secondary"):
                        st.session_state.confirmar_eliminar = viaje_sel; st.rerun()

        if st.session_state.editando_id is not None:
            eid = st.session_state.editando_id
            viaje_edit = db.obtener_por_id(eid)
            if viaje_edit is not None:
                st.divider()
                st.subheader(f"✏️ Editando viaje ID {eid} | Manifiesto: {viaje_edit.get('manifiesto','—')}")
                lista_clientes_edit = get_lista_clientes(db)
                cliente_actual = viaje_edit['cliente']
                if cliente_actual not in lista_clientes_edit: lista_clientes_edit = [cliente_actual] + lista_clientes_edit
                idx_cliente = lista_clientes_edit.index(cliente_actual)
                lista_conductores_edit = get_lista_conductores(db)
                conductor_actual = viaje_edit['conductor']
                if conductor_actual not in lista_conductores_edit: lista_conductores_edit = [conductor_actual] + lista_conductores_edit
                idx_conductor_edit = lista_conductores_edit.index(conductor_actual)
                idx_placa_edit = PLACAS.index(viaje_edit['placa']) if viaje_edit['placa'] in PLACAS else 0

                with st.form(f"form_editar_{eid}"):
                    col1, col2 = st.columns(2)
                    with col1:
                        fecha_e     = st.date_input("Fecha del viaje", value=pd.to_datetime(viaje_edit['fecha_viaje']).date())
                        placa_e     = st.selectbox("Placa", PLACAS, index=idx_placa_edit)
                        conductor_e = st.selectbox("Conductor", lista_conductores_edit, index=idx_conductor_edit)
                        cliente_e   = st.selectbox("Cliente", lista_clientes_edit, index=idx_cliente)
                        manifiesto_e = st.text_input("Número de manifiesto ✱", value=viaje_edit.get('manifiesto','') or '')
                    with col2:
                        origen_e  = st.text_input("Origen",  value=viaje_edit['origen'])
                        destino_e = st.text_input("Destino", value=viaje_edit['destino'])
                        anticipo_e_txt = st.text_input("Valor del anticipo (COP)", value=fmt(viaje_edit['valor_anticipo']))
                        anticipo_e = limpiar(anticipo_e_txt)
                        if anticipo_e > 0: st.caption(f"💵 {fmt(anticipo_e)} COP")
                        obs_e = st.text_area("Observaciones", value=viaje_edit.get('observaciones','') or '', height=80)
                    col_g, col_c = st.columns(2)
                    with col_g: guardar_edit  = st.form_submit_button("💾 Guardar cambios", type="primary")
                    with col_c: cancelar_edit = st.form_submit_button("✖ Cancelar")
                    if guardar_edit:
                        errores_e = []
                        if not manifiesto_e.strip(): errores_e.append("Manifiesto obligatorio")
                        if not origen_e.strip():     errores_e.append("Origen obligatorio")
                        if not destino_e.strip():    errores_e.append("Destino obligatorio")
                        if anticipo_e <= 0:          errores_e.append("Anticipo debe ser mayor a 0")
                        if errores_e:
                            for err in errores_e: st.error(f"⚠️ {err}")
                        else:
                            ok = db.editar_viaje(eid, {
                                'fecha_viaje': fecha_e, 'placa': placa_e,
                                'conductor': conductor_e.strip().upper(),
                                'cliente': cliente_e.strip().upper(),
                                'origen': origen_e.strip().upper(),
                                'destino': destino_e.strip().upper(),
                                'valor_anticipo': anticipo_e,
                                'observaciones': obs_e.strip(),
                                'manifiesto': manifiesto_e.strip()
                            })
                            if ok:
                                st.success(f"✅ Viaje ID {eid} actualizado."); st.session_state.editando_id = None; st.rerun()
                    if cancelar_edit:
                        st.session_state.editando_id = None; st.rerun()

    # ==================== TAB 4: VACACIONES ====================
    with tab_vac:
        st.header("🏖️ Gestión de Vacaciones de Conductores")
        st.markdown("En Colombia, cada trabajador acumula **15 días de vacaciones por año trabajado** (Art. 186 CST).")

        lista_conductores_vac = get_lista_conductores(db)
        df_info_todos         = db.obtener_todos_info_conductores()
        df_vac_todos          = db.obtener_vacaciones()

        # ---- Sub-tabs vacaciones ----
        v_tab1, v_tab2, v_tab3 = st.tabs(["📊 Resumen general", "📝 Registrar vacación", "⚙️ Fecha de ingreso"])

        with v_tab1:
            st.subheader("Estado de vacaciones por conductor")

            # Filtro
            col_f1v, col_f2v = st.columns([2, 2])
            with col_f1v:
                filtro_cond_vac = st.selectbox("Filtrar por conductor", ["Todos"] + lista_conductores_vac, key="vac_filtro_cond")
            with col_f2v:
                filtro_estado_vac = st.selectbox("Estado", ["Todos","Con vacaciones vencidas","Al día","Sin fecha ingreso"], key="vac_filtro_estado")

            conductores_mostrar = lista_conductores_vac if filtro_cond_vac == "Todos" else [filtro_cond_vac]

            resumen_rows = []
            for cond in conductores_mostrar:
                calculo = calcular_vacaciones(cond, df_info_todos, df_vac_todos)
                if calculo["dias_pend"] is None:
                    estado_v = "Sin fecha ingreso"
                elif calculo["dias_pend"] > 15:
                    estado_v = "Vencidas"
                elif calculo["dias_pend"] > 0:
                    estado_v = "Pendientes"
                else:
                    estado_v = "Al día"
                resumen_rows.append({
                    "conductor": cond,
                    "fecha_ingreso": str(calculo["fecha_ingreso"]) if calculo["fecha_ingreso"] else "—",
                    "anios": calculo["anios"],
                    "dias_acum": calculo["dias_acum"],
                    "dias_tomados": calculo["dias_tomados"],
                    "dias_pend": calculo["dias_pend"],
                    "prox_aniv": str(calculo["prox_aniv"]) if calculo["prox_aniv"] else "—",
                    "dias_para_prox": calculo["dias_para_prox"],
                    "estado_v": estado_v,
                })

            df_resumen = pd.DataFrame(resumen_rows)

            if filtro_estado_vac == "Con vacaciones vencidas":
                df_resumen = df_resumen[df_resumen["estado_v"] == "Vencidas"]
            elif filtro_estado_vac == "Al día":
                df_resumen = df_resumen[df_resumen["estado_v"].isin(["Al día"])]
            elif filtro_estado_vac == "Sin fecha ingreso":
                df_resumen = df_resumen[df_resumen["estado_v"] == "Sin fecha ingreso"]

            # Métricas rápidas
            col_mv1, col_mv2, col_mv3, col_mv4 = st.columns(4)
            total_cond    = len(df_resumen)
            con_vencidas  = (df_resumen["estado_v"] == "Vencidas").sum()
            pendientes_v  = (df_resumen["estado_v"] == "Pendientes").sum()
            sin_fecha_v   = (df_resumen["estado_v"] == "Sin fecha ingreso").sum()
            col_mv1.metric("Total conductores", total_cond)
            col_mv2.metric("🔴 Con vencidas",    con_vencidas)
            col_mv3.metric("🟡 Pendientes",      pendientes_v)
            col_mv4.metric("⚪ Sin fecha",        sin_fecha_v)

            if con_vencidas > 0:
                st.error(f"🚨 **{con_vencidas} conductor(es)** tienen vacaciones vencidas (más de 15 días acumulados sin tomar).")

            st.divider()

            # Tabla expandible por conductor
            for _, r in df_resumen.iterrows():
                cond = r["conductor"]
                estado_v = r["estado_v"]
                icono = "🔴" if estado_v == "Vencidas" else ("🟡" if estado_v == "Pendientes" else ("🟢" if estado_v == "Al día" else "⚪"))
                label_v = (f"{icono} {cond}  |  Acumulados: {r['dias_acum'] if r['dias_acum'] is not None else '—'} días  |  "
                           f"Tomados: {r['dias_tomados']} días  |  Pendientes: {r['dias_pend'] if r['dias_pend'] is not None else '—'} días  |  "
                           f"Prox. aniversario: {r['prox_aniv']}")

                with st.expander(label_v):
                    col_va, col_vb = st.columns([2, 2])
                    with col_va:
                        st.markdown(f"**Estado: {estado_v}**")
                        if r["fecha_ingreso"] != "—":
                            st.write(f"📅 Fecha de ingreso: **{r['fecha_ingreso']}**")
                            st.write(f"⏱️ Antigüedad: **{r['anios']} años**")
                            st.write(f"📊 Días acumulados: **{r['dias_acum']}**")
                            st.write(f"✅ Días tomados: **{r['dias_tomados']}**")
                            pend_color = r['dias_pend']
                            if pend_color is not None:
                                if pend_color > 15: st.error(f"⚠️ Días pendientes/vencidos: **{pend_color}**")
                                elif pend_color > 0: st.warning(f"📋 Días pendientes: **{pend_color}**")
                                else: st.success("✅ Vacaciones al día")
                            if r["dias_para_prox"] is not None:
                                st.write(f"📆 Próx. aniversario: **{r['prox_aniv']}** (en {r['dias_para_prox']} días)")
                        else:
                            st.warning("⚠️ No se ha registrado la fecha de ingreso. Regístrala en la pestaña **⚙️ Fecha de ingreso**.")

                    with col_vb:
                        st.markdown("**Historial de vacaciones tomadas:**")
                        df_vac_cond = df_vac_todos[df_vac_todos["conductor"] == cond] if not df_vac_todos.empty else pd.DataFrame()
                        if df_vac_cond.empty:
                            st.info("No hay vacaciones registradas para este conductor.")
                        else:
                            for _, vrow in df_vac_cond.sort_values("fecha_inicio", ascending=False).iterrows():
                                cols_vac = st.columns([3, 1])
                                with cols_vac[0]:
                                    st.write(f"📆 {vrow['fecha_inicio']} → {vrow['fecha_fin']} | **{vrow['dias']} días** | {vrow.get('observaciones','') or '—'}")
                                with cols_vac[1]:
                                    if st.session_state.confirmar_eliminar_vac == vrow['id']:
                                        st.warning("¿Eliminar?")
                                        c_s, c_n = st.columns(2)
                                        with c_s:
                                            if st.button("Sí", key=f"si_vac_{vrow['id']}"):
                                                db.eliminar_vacacion(vrow['id'])
                                                st.session_state.confirmar_eliminar_vac = None
                                                st.success("Registro eliminado."); st.rerun()
                                        with c_n:
                                            if st.button("No", key=f"no_vac_{vrow['id']}"):
                                                st.session_state.confirmar_eliminar_vac = None; st.rerun()
                                    else:
                                        if st.button("🗑️", key=f"del_vac_{vrow['id']}", help="Eliminar registro"):
                                            st.session_state.confirmar_eliminar_vac = vrow['id']; st.rerun()

            st.divider()
            # Exportar Excel vacaciones
            col_exp_v1, col_exp_v2 = st.columns([3, 1])
            with col_exp_v2:
                excel_vac = generar_excel_vacaciones(df_info_todos, df_vac_todos, lista_conductores_vac)
                st.download_button(label="📥 Exportar vacaciones a Excel", data=excel_vac,
                    file_name=f"vacaciones_{hora_colombia().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")

        with v_tab2:
            st.subheader("Registrar período de vacaciones tomado")
            with st.form("form_vacacion", clear_on_submit=True):
                col1v, col2v = st.columns(2)
                with col1v:
                    cond_vac  = st.selectbox("Conductor", lista_conductores_vac, key="vac_cond_reg")
                    fi_vac    = st.date_input("Fecha inicio vacaciones", value=datetime.today())
                    ff_vac    = st.date_input("Fecha fin vacaciones",    value=datetime.today())
                with col2v:
                    # Calcular días automáticamente
                    dias_auto = max(1, (ff_vac - fi_vac).days + 1) if ff_vac >= fi_vac else 1
                    st.metric("Días calculados", dias_auto)
                    dias_vac  = st.number_input("Días de vacaciones (editable)", min_value=1, max_value=60, value=dias_auto)
                    reg_por_v = st.text_input("Registrado por", placeholder="Tu nombre completo")
                    obs_vac   = st.text_area("Observaciones", height=80)

                submitted_vac = st.form_submit_button("💾 Registrar Vacaciones", type="primary")
                if submitted_vac:
                    if not cond_vac:
                        st.error("⚠️ Selecciona un conductor.")
                    elif not reg_por_v.strip():
                        st.error("⚠️ Ingresa tu nombre para registrar.")
                    elif ff_vac < fi_vac:
                        st.error("⚠️ La fecha fin no puede ser anterior a la fecha inicio.")
                    else:
                        nuevo_id_v = db.registrar_vacacion({
                            'conductor': cond_vac, 'fecha_inicio': fi_vac, 'fecha_fin': ff_vac,
                            'dias': dias_vac, 'observaciones': obs_vac.strip(),
                            'registrado_por': reg_por_v.strip()
                        })
                        if nuevo_id_v:
                            st.success(f"✅ Vacaciones registradas para **{cond_vac}**: {fi_vac} → {ff_vac} ({dias_vac} días)")
                            st.rerun()

        with v_tab3:
            st.subheader("⚙️ Registrar/editar fecha de ingreso por conductor")
            st.markdown("Esta fecha es necesaria para calcular los días acumulados de vacaciones.")

            with st.form("form_fecha_ingreso", clear_on_submit=True):
                col1fi, col2fi = st.columns(2)
                with col1fi:
                    cond_fi     = st.selectbox("Conductor", lista_conductores_vac, key="fi_cond")
                    fecha_ing   = st.date_input("Fecha de ingreso (contratación)", value=datetime.today())
                with col2fi:
                    obs_fi = st.text_area("Observaciones", height=80)

                btn_fi = st.form_submit_button("💾 Guardar fecha de ingreso", type="primary")
                if btn_fi:
                    ok_fi = db.guardar_info_conductor(cond_fi, fecha_ing, obs_fi)
                    if ok_fi:
                        st.success(f"✅ Fecha de ingreso de **{cond_fi}** actualizada a **{fecha_ing}**")
                        st.rerun()

            st.divider()
            st.subheader("Fechas de ingreso registradas")
            df_info_show = db.obtener_todos_info_conductores()
            if df_info_show.empty:
                st.info("No hay fechas de ingreso registradas aún.")
            else:
                for _, irow in df_info_show.iterrows():
                    hoy = hora_colombia().date()
                    fi_d = pd.to_datetime(irow['fecha_ingreso']).date()
                    anios_t = round((hoy - fi_d).days / 365.25, 1)
                    st.write(f"👤 **{irow['conductor']}** — Ingreso: {irow['fecha_ingreso']} — Antigüedad: {anios_t} años  |  {irow.get('observaciones','') or ''}")

    # ==================== TAB 5: PRÉSTAMOS ====================
    with tab_prest:
        st.header("💰 Gestión de Préstamos a Conductores")

        lista_conductores_prest = get_lista_conductores(db)

        p_tab1, p_tab2, p_tab3 = st.tabs(["📊 Resumen y trazabilidad", "➕ Nuevo préstamo", "💳 Registrar pago/descuento"])

        with p_tab1:
            st.subheader("Estado de préstamos")

            col_fp1, col_fp2, col_fp3 = st.columns(3)
            with col_fp1: filtro_cond_p  = st.selectbox("Conductor", ["Todos"] + lista_conductores_prest, key="p_filtro_cond")
            with col_fp2: filtro_estado_p = st.selectbox("Estado", ["Todos","activo","saldado"], key="p_filtro_estado")
            with col_fp3: filtro_fecha_p  = st.date_input("Préstamos desde", value=None, key="p_filtro_fecha")

            cond_p_q = None if filtro_cond_p == "Todos" else filtro_cond_p
            est_p_q  = None if filtro_estado_p == "Todos" else filtro_estado_p
            df_prestamos_all = db.obtener_prestamos(conductor=cond_p_q, estado=est_p_q)
            df_pagos_all     = db.obtener_pagos()

            if filtro_fecha_p and not df_prestamos_all.empty:
                df_prestamos_all = df_prestamos_all[pd.to_datetime(df_prestamos_all["fecha_prestamo"]).dt.date >= filtro_fecha_p]

            if df_prestamos_all.empty:
                st.info("No se encontraron préstamos con los filtros aplicados.")
            else:
                # Calcular saldos
                total_prestado = int(df_prestamos_all["monto_total"].sum())
                total_pagado_g = 0
                total_saldo_g  = 0
                activos_g      = 0
                saldados_g     = 0
                for _, pr in df_prestamos_all.iterrows():
                    pagado, saldo = calcular_saldo_prestamo(pr['id'], pr['monto_total'], df_pagos_all)
                    total_pagado_g += pagado
                    total_saldo_g  += saldo
                    if pr['estado'] == 'activo': activos_g  += 1
                    else:                        saldados_g += 1

                col_pm1, col_pm2, col_pm3, col_pm4 = st.columns(4)
                col_pm1.metric("Total prestado",  f"${fmt(total_prestado)}")
                col_pm2.metric("Total pagado",    f"${fmt(total_pagado_g)}")
                col_pm3.metric("Saldo pendiente", f"${fmt(total_saldo_g)}")
                col_pm4.metric("Activos / Saldados", f"{activos_g} / {saldados_g}")

                if total_saldo_g > 0:
                    st.warning(f"💰 Saldo total pendiente de cobro: **${fmt(total_saldo_g)} COP**")

                st.divider()

                # Expanders por préstamo
                for _, pr in df_prestamos_all.iterrows():
                    pagado, saldo = calcular_saldo_prestamo(pr['id'], pr['monto_total'], df_pagos_all)
                    pct = round(pagado / pr['monto_total'] * 100) if pr['monto_total'] > 0 else 0
                    paz_salvo = pr['estado'] == 'saldado' or saldo == 0
                    icono_p = "✅" if paz_salvo else "🔴"
                    label_p = (f"{icono_p} ID {pr['id']} | {pr['conductor']} | "
                               f"Préstamo: ${fmt(pr['monto_total'])} | Pagado: ${fmt(pagado)} | "
                               f"Saldo: ${fmt(saldo)} | {pct}% | {pr['estado'].upper()}")

                    with st.expander(label_p):
                        col_pa, col_pb = st.columns([2, 2])
                        with col_pa:
                            st.markdown(f"**Conductor:** {pr['conductor']}")
                            st.write(f"📅 Fecha préstamo: {pr['fecha_prestamo']}")
                            st.write(f"💰 Monto total: **${fmt(pr['monto_total'])} COP**")
                            st.write(f"✅ Total pagado: **${fmt(pagado)} COP**")
                            if saldo > 0:
                                st.error(f"📋 Saldo pendiente: **${fmt(saldo)} COP** ({100-pct}% restante)")
                            else:
                                st.success("✅ PAZ Y SALVO — Préstamo cancelado en su totalidad")
                            if pr.get('motivo'): st.write(f"📝 Motivo: {pr['motivo']}")
                            if pr.get('observaciones'): st.write(f"📝 Obs: {pr['observaciones']}")

                            # Barra de progreso
                            st.progress(min(pct, 100) / 100, text=f"Progreso de pago: {pct}%")

                            # Botón marcar paz y salvo manualmente
                            if not paz_salvo:
                                col_btna, col_btnb = st.columns(2)
                                with col_btna:
                                    if st.button("✅ Marcar Paz y Salvo", key=f"paz_{pr['id']}", type="primary"):
                                        db.actualizar_estado_prestamo(pr['id'], 'saldado')
                                        st.success(f"✅ Préstamo ID {pr['id']} marcado como Paz y Salvo."); st.rerun()
                                with col_btnb:
                                    if st.session_state.confirmar_eliminar_prestamo == pr['id']:
                                        st.warning("¿Eliminar préstamo y todos sus pagos?")
                                        c_s2, c_n2 = st.columns(2)
                                        with c_s2:
                                            if st.button("Sí", key=f"si_prest_{pr['id']}"):
                                                db.eliminar_prestamo(pr['id'])
                                                st.session_state.confirmar_eliminar_prestamo = None
                                                st.success("Préstamo eliminado."); st.rerun()
                                        with c_n2:
                                            if st.button("No", key=f"no_prest_{pr['id']}"):
                                                st.session_state.confirmar_eliminar_prestamo = None; st.rerun()
                                    else:
                                        if st.button("🗑️ Eliminar", key=f"del_prest_{pr['id']}"):
                                            st.session_state.confirmar_eliminar_prestamo = pr['id']; st.rerun()
                            else:
                                if st.button("↩️ Reabrir préstamo", key=f"reabrir_{pr['id']}"):
                                    db.actualizar_estado_prestamo(pr['id'], 'activo'); st.rerun()

                        with col_pb:
                            st.markdown("**Historial de pagos / descuentos:**")
                            df_pagos_p = df_pagos_all[df_pagos_all["prestamo_id"] == pr['id']] if not df_pagos_all.empty else pd.DataFrame()
                            if df_pagos_p.empty:
                                st.info("No hay pagos registrados aún.")
                            else:
                                saldo_acum = int(pr['monto_total'])
                                for _, pg in df_pagos_p.sort_values("fecha_pago").iterrows():
                                    saldo_acum -= int(pg['monto_pago'])
                                    col_pgr = st.columns([3, 1])
                                    with col_pgr[0]:
                                        st.write(f"💳 {pg['fecha_pago']} — Descuento: **${fmt(pg['monto_pago'])}** — Saldo restante: **${fmt(max(0,saldo_acum))}**"
                                                 + (f" | {pg['observaciones']}" if pg.get('observaciones') else ""))
                                    with col_pgr[1]:
                                        if st.session_state.confirmar_eliminar_pago == pg['id']:
                                            st.warning("¿Eliminar?")
                                            c_s3, c_n3 = st.columns(2)
                                            with c_s3:
                                                if st.button("Sí", key=f"si_pago_{pg['id']}"):
                                                    db.eliminar_pago(pg['id'])
                                                    st.session_state.confirmar_eliminar_pago = None
                                                    st.success("Pago eliminado."); st.rerun()
                                            with c_n3:
                                                if st.button("No", key=f"no_pago_{pg['id']}"):
                                                    st.session_state.confirmar_eliminar_pago = None; st.rerun()
                                        else:
                                            if st.button("🗑️", key=f"del_pago_{pg['id']}", help="Eliminar pago"):
                                                st.session_state.confirmar_eliminar_pago = pg['id']; st.rerun()

                st.divider()
                col_exp_p1, col_exp_p2 = st.columns([3, 1])
                with col_exp_p2:
                    df_pagos_export = db.obtener_pagos()
                    excel_p = generar_excel_prestamos(df_prestamos_all, df_pagos_export)
                    st.download_button(label="📥 Exportar a Excel", data=excel_p,
                        file_name=f"prestamos_{hora_colombia().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")

        with p_tab2:
            st.subheader("Registrar nuevo préstamo")
            with st.form("form_prestamo", clear_on_submit=True):
                col1p, col2p = st.columns(2)
                with col1p:
                    cond_nuevo_p  = st.selectbox("Conductor", lista_conductores_prest, key="p_cond_nuevo")
                    fecha_prest   = st.date_input("Fecha del préstamo", value=datetime.today())
                    monto_prest_txt = st.text_input("Monto del préstamo (COP)", placeholder="Ej: 500.000")
                    monto_prest   = limpiar(monto_prest_txt)
                    if monto_prest > 0: st.caption(f"💵 {fmt(monto_prest)} COP")
                with col2p:
                    motivo_prest  = st.text_input("Motivo del préstamo", placeholder="Ej: Urgencia médica, anticipo salario...")
                    obs_prest     = st.text_area("Observaciones", height=80)

                btn_prest = st.form_submit_button("💾 Registrar Préstamo", type="primary")
                if btn_prest:
                    if monto_prest <= 0:
                        st.error("⚠️ El monto debe ser mayor a 0.")
                    else:
                        nid_p = db.registrar_prestamo({
                            'conductor': cond_nuevo_p,
                            'monto_total': monto_prest,
                            'fecha_prestamo': fecha_prest,
                            'motivo': motivo_prest.strip(),
                            'observaciones': obs_prest.strip()
                        })
                        if nid_p:
                            st.success(f"✅ Préstamo registrado (ID: {nid_p}) para **{cond_nuevo_p}** por **${fmt(monto_prest)} COP**")
                            st.rerun()

        with p_tab3:
            st.subheader("Registrar pago / descuento a un préstamo")
            st.markdown("Selecciona el préstamo activo del conductor al que deseas aplicar un descuento.")

            # Mostrar solo préstamos activos con saldo
            df_activos = db.obtener_prestamos(estado="activo")
            df_pagos_check = db.obtener_pagos()

            if df_activos.empty:
                st.success("✅ No hay préstamos activos en este momento.")
            else:
                opciones_prestamos = []
                for _, pr in df_activos.iterrows():
                    pagado, saldo = calcular_saldo_prestamo(pr['id'], pr['monto_total'], df_pagos_check)
                    if saldo > 0:
                        opciones_prestamos.append({
                            "id": pr['id'],
                            "label": f"ID {pr['id']} | {pr['conductor']} | Saldo: ${fmt(saldo)} de ${fmt(pr['monto_total'])}",
                            "saldo": saldo,
                            "conductor": pr['conductor'],
                            "monto_total": pr['monto_total']
                        })

                if not opciones_prestamos:
                    st.success("✅ Todos los préstamos activos están saldados. Marca los como Paz y Salvo en la pestaña de Resumen.")
                else:
                    with st.form("form_pago", clear_on_submit=True):
                        prestamo_sel_idx = st.selectbox(
                            "Selecciona el préstamo",
                            range(len(opciones_prestamos)),
                            format_func=lambda i: opciones_prestamos[i]["label"],
                            key="p_sel_pago"
                        )
                        op_sel = opciones_prestamos[prestamo_sel_idx]
                        st.info(f"💰 Saldo actual del préstamo seleccionado: **${fmt(op_sel['saldo'])} COP**")

                        col1pg, col2pg = st.columns(2)
                        with col1pg:
                            fecha_pago    = st.date_input("Fecha del descuento", value=datetime.today())
                            monto_pago_txt = st.text_input("Monto del descuento (COP)", placeholder="Ej: 50.000")
                            monto_pago    = limpiar(monto_pago_txt)
                            if monto_pago > 0:
                                saldo_restante = max(0, op_sel['saldo'] - monto_pago)
                                st.caption(f"💵 Descuento: {fmt(monto_pago)} COP — Saldo restante: **{fmt(saldo_restante)} COP**")
                        with col2pg:
                            reg_por_pg = st.text_input("Registrado por", placeholder="Tu nombre completo")
                            obs_pg     = st.text_area("Observaciones", height=80, placeholder="Nómina julio, descuento quincenal...")

                        btn_pago = st.form_submit_button("💳 Registrar Descuento", type="primary")
                        if btn_pago:
                            if monto_pago <= 0:
                                st.error("⚠️ El monto del descuento debe ser mayor a 0.")
                            elif monto_pago > op_sel['saldo']:
                                st.error(f"⚠️ El descuento (${fmt(monto_pago)}) supera el saldo (${fmt(op_sel['saldo'])}). Ajusta el valor.")
                            elif not reg_por_pg.strip():
                                st.error("⚠️ Ingresa tu nombre para registrar.")
                            else:
                                nid_pg = db.registrar_pago({
                                    'prestamo_id': op_sel['id'],
                                    'monto_pago': monto_pago,
                                    'fecha_pago': fecha_pago,
                                    'observaciones': obs_pg.strip(),
                                    'registrado_por': reg_por_pg.strip()
                                })
                                if nid_pg:
                                    nuevo_saldo = max(0, op_sel['saldo'] - monto_pago)
                                    st.success(f"✅ Descuento de **${fmt(monto_pago)} COP** registrado para **{op_sel['conductor']}**. Saldo restante: **${fmt(nuevo_saldo)} COP**")
                                    # Auto marcar paz y salvo
                                    if nuevo_saldo == 0:
                                        db.actualizar_estado_prestamo(op_sel['id'], 'saldado')
                                        st.success(f"🎉 ¡**{op_sel['conductor']}** quedó **PAZ Y SALVO** con este préstamo!")
                                    st.rerun()

    # ==================== TAB 6: CLIENTES ====================
    with tab_clientes:
        st.header("🏢 Gestión de Clientes")
        st.subheader("Clientes predeterminados")
        cols = st.columns(len(CLIENTES_DEFAULT))
        for i, c_def in enumerate(CLIENTES_DEFAULT):
            with cols[i]: st.info(c_def)
        st.divider()
        st.subheader("Agregar cliente nuevo")
        with st.form("form_nuevo_cliente", clear_on_submit=True):
            nuevo_cliente = st.text_input("Nombre del cliente", placeholder="Ej: LOGÍSTICA DEL NORTE")
            if st.form_submit_button("➕ Agregar Cliente", type="primary"):
                if not nuevo_cliente.strip():
                    st.error("⚠️ El nombre no puede estar vacío.")
                elif nuevo_cliente.strip().upper() in [c.upper() for c in CLIENTES_DEFAULT]:
                    st.warning("⚠️ Ya existe en la lista predeterminada.")
                else:
                    ok = db.agregar_cliente(nuevo_cliente.strip())
                    if ok: st.success(f"✅ Cliente **{nuevo_cliente.strip().upper()}** agregado."); st.rerun()
                    else: st.error("❌ Ya existe o hubo un error.")
        st.divider()
        st.subheader("Clientes adicionales registrados")
        df_extras = db.obtener_clientes_extra()
        if df_extras.empty:
            st.info("No hay clientes adicionales aún.")
        else:
            for _, row in df_extras.iterrows():
                col_n, col_f, col_b = st.columns([3, 2, 1])
                with col_n: st.write(f"**{row['nombre']}**")
                with col_f: st.write(f"Registrado: {str(row['fecha_registro'])[:16]}")
                with col_b:
                    if st.session_state.confirmar_eliminar_cliente == row['id']:
                        st.warning("¿Eliminar?")
                        c_si, c_no = st.columns(2)
                        with c_si:
                            if st.button("Sí", key=f"si_cli_{row['id']}"):
                                db.eliminar_cliente(row['id']); st.session_state.confirmar_eliminar_cliente = None
                                st.success("Eliminado."); st.rerun()
                        with c_no:
                            if st.button("No", key=f"no_cli_{row['id']}"):
                                st.session_state.confirmar_eliminar_cliente = None; st.rerun()
                    else:
                        if st.button("🗑️", key=f"del_cli_{row['id']}"):
                            st.session_state.confirmar_eliminar_cliente = row['id']; st.rerun()
        st.divider()
        st.subheader("Lista completa")
        for c in get_lista_clientes(db): st.write(f"• {c}")

    # ==================== TAB 7: CONDUCTORES ====================
    with tab_conductores:
        st.header("👤 Gestión de Conductores")
        st.markdown("Agrega, edita o elimina conductores adicionales.")

        st.subheader("Conductores predeterminados")
        cols_def = st.columns(4)
        for i, c_def in enumerate(sorted(CONDUCTORES_DEFAULT)):
            with cols_def[i % 4]: st.info(c_def)

        st.divider()
        st.subheader("Agregar conductor nuevo")
        with st.form("form_nuevo_conductor", clear_on_submit=True):
            nuevo_conductor = st.text_input("Nombre del conductor", placeholder="Ej: JUAN PABLO GOMEZ")
            if st.form_submit_button("➕ Agregar Conductor", type="primary"):
                if not nuevo_conductor.strip():
                    st.error("⚠️ El nombre no puede estar vacío.")
                elif nuevo_conductor.strip().upper() in [c.upper() for c in CONDUCTORES_DEFAULT]:
                    st.warning("⚠️ Ya existe en la lista predeterminada.")
                else:
                    ok = db.agregar_conductor(nuevo_conductor.strip())
                    if ok: st.success(f"✅ Conductor **{nuevo_conductor.strip().upper()}** agregado."); st.rerun()
                    else: st.error("❌ Ya existe o hubo un error.")

        st.divider()
        st.subheader("Conductores adicionales registrados")
        df_conductores = db.obtener_conductores_extra()
        if df_conductores.empty:
            st.info("No hay conductores adicionales registrados aún.")
        else:
            for _, row in df_conductores.iterrows():
                col_nombre, col_fecha, col_edit, col_del = st.columns([3, 2, 1, 1])
                with col_nombre:
                    if st.session_state.editando_conductor_id == row['id']:
                        nombre_editado = st.text_input("Nuevo nombre", value=row['nombre'],
                            key=f"edit_input_{row['id']}", label_visibility="collapsed")
                    else:
                        st.write(f"**{row['nombre']}**")
                with col_fecha:
                    st.write(f"Registrado: {str(row['fecha_registro'])[:16]}")
                with col_edit:
                    if st.session_state.editando_conductor_id == row['id']:
                        if st.button("💾", key=f"save_cond_{row['id']}", help="Guardar cambios"):
                            if nombre_editado.strip():
                                ok = db.editar_conductor(row['id'], nombre_editado.strip())
                                if ok:
                                    st.success(f"✅ Actualizado a **{nombre_editado.strip().upper()}**")
                                    st.session_state.editando_conductor_id = None; st.rerun()
                            else: st.error("El nombre no puede estar vacío.")
                    else:
                        if st.button("✏️", key=f"edit_cond_{row['id']}", help="Editar nombre"):
                            st.session_state.editando_conductor_id = row['id']
                            st.session_state.confirmar_eliminar_conductor = None; st.rerun()
                with col_del:
                    if st.session_state.editando_conductor_id == row['id']:
                        if st.button("✖", key=f"cancel_cond_{row['id']}", help="Cancelar edición"):
                            st.session_state.editando_conductor_id = None; st.rerun()
                    elif st.session_state.confirmar_eliminar_conductor == row['id']:
                        st.warning("¿Eliminar?")
                        c_si2, c_no2 = st.columns(2)
                        with c_si2:
                            if st.button("Sí", key=f"si_cond_{row['id']}"):
                                db.eliminar_conductor(row['id']); st.session_state.confirmar_eliminar_conductor = None
                                st.success("Conductor eliminado."); st.rerun()
                        with c_no2:
                            if st.button("No", key=f"no_cond_{row['id']}"):
                                st.session_state.confirmar_eliminar_conductor = None; st.rerun()
                    else:
                        if st.button("🗑️", key=f"del_cond_{row['id']}", help="Eliminar conductor"):
                            st.session_state.confirmar_eliminar_conductor = row['id']
                            st.session_state.editando_conductor_id = None; st.rerun()

        st.divider()
        st.subheader("Lista completa de conductores")
        todos_conductores = get_lista_conductores(db)
        cols_todos = st.columns(3)
        for i, c in enumerate(todos_conductores):
            with cols_todos[i % 3]: st.write(f"• {c}")


if __name__ == "__main__":
    main()
